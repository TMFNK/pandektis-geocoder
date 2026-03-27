#!/usr/bin/env python3
"""
Greece-GEO: Pandektis Settlement Renamings Scraper & Geocoder

Extracts ~4,413 settlement renaming records from the National Documentation
Centre of Greece (Pandektis), geocodes them, and exports as GeoJSON + Excel.

Usage:
    python3 scrape_pandektis.py [--skip-scrape] [--skip-geocode] [--max-entries N]
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_URL = "https://pandektis.ekt.gr"
BROWSE_URL = f"{BASE_URL}/pandektis/handle/10442/4968/browse"
RPP = 100  # max results per page on the browse endpoint
REQUEST_DELAY = 0.4  # seconds between requests to Pandektis (polite)
GEOCODE_DELAY = 1.1  # seconds between Nominatim calls (policy: 1 req/s)
USER_AGENT = "Greece-GEO-Research-Scraper/1.0 (academic; settlement-geocoding)"
OUTPUT_DIR = Path(__file__).parent
PROGRESS_FILE = OUTPUT_DIR / "progress.json"
FAILED_FILE = OUTPUT_DIR / "failed_geocoding.csv"

# Prefecture name normalization: map ALL-CAPS English → Title Case for geocoding
PREFECTURE_OVERRIDES = {
    "ΑΤΤΙΚΗΣ": "Attica",
    "ΑΤΤΙΚΙ": "Attica",
    "ΘΕΣΣΑΛΟΝΙΚΗΣ": "Thessaloniki",
    "ΑΧΑΪΑΣ": "Achaia",
    "ΗΡΑΚΛΕΙΟΥ": "Heraklion",
    "ΛΑΡΙΣΑΣ": "Larissa",
    "ΜΑΚΕΔΟΝΙΑΣ": "Macedonia",
}

# ---------------------------------------------------------------------------
# Phase 1: Scrape all entry URLs from browse pages
# ---------------------------------------------------------------------------


def collect_entry_urls(session: requests.Session) -> list[dict]:
    """Paginate browse pages and collect all entry URLs + display titles."""
    entries = []
    seen_urls = set()
    offset = 0
    page_num = 0

    while True:
        page_num += 1
        params = {
            "type": "title",
            "sort_by": 1,
            "order": "ASC",
            "rpp": RPP,
            "etal": -1,
            "offset": offset,
        }
        print(
            f"  [Browse] Fetching page {page_num} (offset={offset})...",
            end=" ",
            flush=True,
        )
        resp = session.get(BROWSE_URL, params=params, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        links = soup.select("a.ekt_browse_no_image")
        if not links:
            print("no more entries.")
            break

        new_count = 0
        all_duplicates = True
        for link in links:
            href = link.get("href", "")
            title = link.get_text(strip=True)
            if href:
                full_url = BASE_URL + href if href.startswith("/") else href
                if full_url not in seen_urls:
                    seen_urls.add(full_url)
                    entries.append({"url": full_url, "display_title": title})
                    new_count += 1
                    all_duplicates = False

        if all_duplicates and new_count == 0:
            print("all duplicates — dataset complete.")
            break

        print(f"found {len(links)} entries, {new_count} new (total: {len(entries)})")
        offset += RPP
        time.sleep(REQUEST_DELAY)

    return entries


# ---------------------------------------------------------------------------
# Phase 2: Parse individual entry pages
# ---------------------------------------------------------------------------


def parse_entry_page(
    session: requests.Session, url: str, max_retries: int = 3
) -> dict | None:
    """Fetch a single entry page and extract all metadata fields. Retries on 500 errors."""
    for attempt in range(max_retries):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code >= 500:
                wait = 2 ** (attempt + 1)  # 2, 4, 8 seconds
                print(
                    f"retry {attempt + 1}/{max_retries} (HTTP {resp.status_code}, wait {wait}s)",
                    end=" ",
                    flush=True,
                )
                time.sleep(wait)
                continue
            resp.raise_for_status()
            break
        except requests.exceptions.ConnectionError:
            if attempt < max_retries - 1:
                time.sleep(3)
                continue
            raise
    else:
        raise Exception(f"Failed after {max_retries} retries (500 errors)")

    soup = BeautifulSoup(resp.text, "lxml")

    # Extract all label-value pairs from ekt_met_curved / ekt_met_curved_metadata
    labels = soup.select("div.ekt_met_curved")
    values = soup.select("div.ekt_met_curved_metadata")

    if len(labels) < 2:
        return None

    # Build a dict of label → value (strip whitespace and &nbsp)
    pairs = {}
    for label, value in zip(labels, values):
        key = (
            label.get_text(strip=True)
            .replace("\xa0", "")
            .replace("&nbsp", "")
            .strip()
            .rstrip(":")
            .strip()
        )
        val = value.get_text(strip=True)
        pairs[key] = val

    # Extract fields using Greek and English label keys
    old_name_gr = pairs.get("Παλαιά ονομασία", "")
    old_name_en = pairs.get("Old name", "")
    new_name_gr = pairs.get("Νέα ονομασία", "")
    new_name_en = pairs.get("New name", "")
    prefecture_en = pairs.get("Prefecture", "")
    province_en = pairs.get("Province", "")
    renaming_date = pairs.get("Date of renaming", "")
    settlement_code = pairs.get("Code of settlement", "")
    official_journal = pairs.get("Official Journal", "")
    municipality_en = pairs.get("Name of Community or Municipality", "")
    municipality_gr = pairs.get("Κοινότητα ή Δήμος", "")

    # Check if this page links to a modern settlement record
    # Some entries have links like "Modern settlement record" in the page
    modern_link = None
    for a_tag in soup.select("a"):
        href = a_tag.get("href", "")
        text = a_tag.get_text(strip=True).lower()
        if "modern" in text or "σύγχρονο" in text or "handle/10442" in href:
            if href != url and "handle/10442" in href and "4968" not in href:
                modern_link = BASE_URL + href if href.startswith("/") else href
                break

    return {
        "old_name_gr": old_name_gr,
        "old_name_en": old_name_en,
        "new_name_gr": new_name_gr,
        "new_name_en": new_name_en,
        "prefecture": prefecture_en,
        "province": province_en,
        "renaming_date": renaming_date,
        "settlement_code": settlement_code,
        "official_journal": official_journal,
        "municipality_en": municipality_en,
        "municipality_gr": municipality_gr,
        "source_url": url,
        "modern_link": modern_link,
    }


def follow_modern_link(session: requests.Session, url: str) -> dict | None:
    """Follow a modern settlement record link to get more accurate data."""
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        labels = soup.select("div.ekt_met_curved")
        values = soup.select("div.ekt_met_curved_metadata")
        pairs = {}
        for label, value in zip(labels, values):
            key = (
                label.get_text(strip=True)
                .replace("\xa0", "")
                .strip()
                .rstrip(":")
                .strip()
            )
            val = value.get_text(strip=True)
            pairs[key] = val
        return pairs
    except Exception:
        return None


def scrape_all_entries(
    session: requests.Session, entry_urls: list[dict], skip_existing: set = None
) -> list[dict]:
    """Scrape metadata from all entry pages."""
    if skip_existing is None:
        skip_existing = set()

    records = []
    total = len(entry_urls)
    failed = []

    for i, entry in enumerate(entry_urls):
        url = entry["url"]
        if url in skip_existing:
            continue

        print(
            f"  [Parse] {i + 1}/{total}: {entry['display_title'][:60]}...",
            end=" ",
            flush=True,
        )
        try:
            record = parse_entry_page(session, url)
            if record is None:
                print("SKIPPED (no data)")
                continue

            # If there's a modern link, follow it for more accurate new_name
            if record.get("modern_link"):
                modern = follow_modern_link(session, record["modern_link"])
                if modern:
                    # Use modern record's new name if available
                    modern_new_en = modern.get("New name", "")
                    modern_new_gr = modern.get("Νέα ονομασία", "")
                    if modern_new_en:
                        record["new_name_en"] = modern_new_en
                    if modern_new_gr:
                        record["new_name_gr"] = modern_new_gr
                    # Also get coordinates from modern record if present
                    for key in modern:
                        if (
                            "latitude" in key.lower()
                            or "longitude" in key.lower()
                            or "γεωγραφ" in key.lower()
                        ):
                            record[f"modern_{key}"] = modern[key]

            records.append(record)
            print("OK")

        except Exception as e:
            print(f"FAILED ({e})")
            failed.append(
                {"url": url, "title": entry["display_title"], "error": str(e)}
            )

        time.sleep(REQUEST_DELAY)

    if failed:
        print(f"\n  {len(failed)} entries failed to parse.")
        with open(
            OUTPUT_DIR / "failed_parsing.csv", "w", newline="", encoding="utf-8"
        ) as f:
            writer = csv.DictWriter(f, fieldnames=["url", "title", "error"])
            writer.writeheader()
            writer.writerows(failed)

    return records


# ---------------------------------------------------------------------------
# Phase 3: Geocode via Nominatim (multi-strategy)
# ---------------------------------------------------------------------------

# Prefecture bounding boxes [min_lat, max_lat, min_lon, max_lon]
# Used to validate that geocoded results land in the right region
PREFECTURE_BOUNDS = {
    "ATTIKI": (37.7, 38.3, 23.3, 24.1),
    "THESSALONIKI": (40.2, 40.9, 22.4, 23.6),
    "KILKIS": (40.7, 41.3, 22.2, 23.2),
    "PIERIA": (39.9, 40.5, 22.1, 22.7),
    "IMATHIA": (40.3, 40.7, 21.8, 22.3),
    "PELLA": (40.5, 41.2, 21.8, 22.2),
    "SERRES": (40.7, 41.4, 23.2, 24.1),
    "CHALKIDIKI": (39.9, 40.7, 23.2, 24.0),
    "DRAMA": (40.9, 41.5, 23.8, 24.6),
    "KAVALA": (40.7, 41.1, 24.0, 24.6),
    "XANTHI": (41.0, 41.4, 24.5, 25.0),
    "RODOPI": (40.9, 41.3, 25.2, 26.0),
    "EVROS": (41.2, 41.8, 25.8, 26.6),
    "LARISSA": (39.3, 40.1, 22.0, 22.8),
    "MAGNISIA": (38.9, 39.7, 22.3, 23.2),
    "KARDITSA": (39.1, 39.5, 21.5, 22.1),
    "TRIKALA": (39.3, 39.9, 21.2, 21.9),
    "GREVENA": (39.8, 40.3, 21.2, 21.8),
    "KOZANI": (40.0, 40.6, 21.4, 22.0),
    "KASTORIA": (40.3, 40.7, 21.1, 21.7),
    "FLORINA": (40.6, 41.0, 21.2, 21.7),
    "IOANNINA": (39.4, 40.0, 20.5, 21.2),
    "THESPROTIA": (39.3, 39.9, 20.1, 20.8),
    "PREVEZA": (38.9, 39.4, 20.5, 21.0),
    "ARTA": (38.9, 39.4, 20.8, 21.3),
    "EVOIA": (38.3, 39.1, 22.9, 24.2),
    "VOIOTIA": (38.1, 38.7, 22.7, 23.5),
    "FTHIOTIS": (38.5, 39.1, 21.9, 23.0),
    "FOKIDA": (38.3, 38.8, 21.8, 22.5),
    "AITOLOAKARNANIA": (38.3, 39.2, 20.9, 21.9),
    "KERKYRA": (39.3, 39.9, 19.6, 20.2),
    "KEFALONIA": (38.0, 38.5, 20.3, 20.9),
    "LEFKAS": (38.5, 38.9, 20.4, 20.8),
    "ZAKYNTHOS": (37.6, 38.0, 20.6, 21.0),
    "KYKLADES": (36.3, 37.5, 24.0, 25.8),
    "IRAKLEIO": (34.9, 35.6, 24.7, 25.7),
    "LASITHI": (34.9, 35.3, 25.4, 26.3),
    "RETHYMNO": (35.1, 35.5, 24.2, 24.9),
    "CHANIA": (35.1, 35.7, 23.5, 24.3),
    "ACHAIA": (37.8, 38.3, 21.6, 22.3),
    "ILIA": (37.5, 38.0, 21.2, 21.9),
    "ARKADIA": (37.2, 37.8, 21.7, 22.4),
    "KORINTHIA": (37.7, 38.2, 22.2, 22.9),
    "ARGOLIDA": (37.3, 37.8, 22.4, 23.2),
    "LAKONIA": (36.3, 37.2, 22.0, 22.8),
    "MESSINIA": (36.6, 37.3, 21.5, 22.2),
    "LIMNOS": (39.7, 40.1, 25.1, 25.5),
    "LESVOS": (38.9, 39.4, 25.9, 26.6),
    "CHIOS": (38.2, 38.6, 25.9, 26.2),
    "SAMOS": (37.6, 37.9, 26.5, 27.1),
    "DODEKANISA": (36.0, 37.5, 26.0, 28.5),
    "MAKEDONIAS": (40.0, 41.5, 21.0, 26.0),
}


def _nominatim_search(query: str, extra_params: dict = None) -> list[dict]:
    """Call Nominatim search API directly. Returns list of results."""
    params = {
        "q": query,
        "format": "json",
        "limit": 3,
        "accept-language": "en",
        "countrycodes": "gr",
    }
    if extra_params:
        params.update(extra_params)
    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params=params,
            headers={"User-Agent": USER_AGENT},
            timeout=10,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception:
        return []


def _validate_bounds(lat: float, lon: float, prefecture: str) -> bool:
    """Check if coordinates fall within Greece and optionally the expected prefecture."""
    # General Greece bounds
    if not (34.0 <= lat <= 42.0 and 19.0 <= lon <= 30.0):
        return False
    # Prefecture-specific bounds if available
    if prefecture in PREFECTURE_BOUNDS:
        min_lat, max_lat, min_lon, max_lon = PREFECTURE_BOUNDS[prefecture]
        # Allow 0.3 degree margin for boundary cases
        margin = 0.3
        if not (
            (min_lat - margin) <= lat <= (max_lat + margin)
            and (min_lon - margin) <= lon <= (max_lon + margin)
        ):
            return False
    return True


# Prefecture name mapping: English ALL-CAPS → Greek
PREFECTURE_EN_TO_GR = {
    "ATTIKI": "Αττικής",
    "THESSALONIKI": "Θεσσαλονίκης",
    "KILKIS": "Κιλκίς",
    "PIERIA": "Πιερίας",
    "IMATHIA": "Ημαθίας",
    "PELLA": "Πέλλας",
    "SERRES": "Σερρών",
    "CHALKIDIKI": "Χαλκιδικής",
    "DRAMA": "Δράμας",
    "KAVALA": "Καβάλας",
    "XANTHI": "Ξάνθης",
    "RODOPI": "Ροδόπης",
    "EVROS": "Έβρου",
    "LARISSA": "Λάρισας",
    "MAGNISIA": "Μαγνησίας",
    "KARDITSA": "Καρδίτσας",
    "TRIKALA": "Τρικάλων",
    "GREVENA": "Γρεβενών",
    "KOZANI": "Κοζάνης",
    "KASTORIA": "Καστοριάς",
    "FLORINA": "Φλώρινας",
    "IOANNINA": "Ιωαννίνων",
    "THESPROTIA": "Θεσπρωτίας",
    "PREVEZA": "Πρέβεζας",
    "ARTA": "Άρτας",
    "EVOIA": "Εύβοιας",
    "VOIOTIA": "Βοιωτίας",
    "FTHIOTIS": "Φθιώτιδας",
    "FOKIDA": "Φωκίδας",
    "AITOLOAKARNANIA": "Αιτωλοακαρνανίας",
    "KERKYRA": "Κέρκυρας",
    "KEFALONIA": "Κεφαλλονίας",
    "LEFKAS": "Λευκάδας",
    "ZAKYNTHOS": "Ζακύνθου",
    "KYKLADES": "Κυκλάδων",
    "IRAKLEIO": "Ηρακλείου",
    "LASITHI": "Λασιθίου",
    "RETHYMNO": "Ρεθύμνης",
    "CHANIA": "Χανίων",
    "ACHAIA": "Αχαΐας",
    "ILIA": "Ηλείας",
    "ARKADIA": "Αρκαδίας",
    "KORINTHIA": "Κορινθίας",
    "ARGOLIDA": "Αργολίδας",
    "LAKONIA": "Λακωνίας",
    "MESSINIA": "Μεσσηνίας",
    "LIMNOS": "Λήμνου",
    "LESVOS": "Λέσβου",
    "CHIOS": "Χίου",
    "SAMOS": "Σάμου",
    "DODEKANISA": "Δωδεκανήσου",
}


def _prefecture_to_greek(prefecture_en: str) -> str:
    """Convert English ALL-CAPS prefecture name to Greek genitive."""
    return PREFECTURE_EN_TO_GR.get(prefecture_en.upper(), "")


def _greek_genitive_to_nominative(genitive: str) -> str:
    """
    Attempt to convert a Greek genitive noun to nominative form.
    This is heuristic — covers common patterns for place names.
    """
    g = genitive.strip()
    # Common genitive → nominative endings
    # -ου (masculine/neuter 2nd declension) → -ος or -ο
    if g.endswith("ου") and len(g) > 3:
        return g[:-2] + "ος"
    # -ίου → -ι or -ία
    if g.endswith("ίου") and len(g) > 4:
        return g[:-3] + "ία"
    # -ας (feminine 1st declension) → -α
    if g.endswith("ας") and len(g) > 3:
        return g[:-2]
    # -ων (feminine/plural genitive) → keep as is (often same)
    # -ης → -η or -ης
    if g.endswith("ης") and len(g) > 3:
        return g[:-1]
    return g


def geocode_settlement(
    record: dict,
) -> tuple[float | None, float | None, str]:
    """
    Multi-strategy geocoding with Greek name and municipality fallbacks.
    Returns (lat, lon, query_used) or (None, None, query_failed).
    """
    name_en = record.get("new_name_en", "").strip()
    name_gr = record.get("new_name_gr", "").strip()
    prefecture = record.get("prefecture", "").strip()
    municipality_en = record.get("municipality_en", "").strip()
    municipality_gr = record.get("municipality_gr", "").strip()

    if not name_en and not name_gr:
        return None, None, ""

    # Build a prioritized query chain
    # 1. English name + prefecture (most specific)
    # 2. Greek name + prefecture (handles transliteration gaps)
    # 3. Municipality (EN) + prefecture (fallback to municipality coordinates)
    # 4. Municipality (GR) + Greece
    # 5. Greek name alone
    # 6. English name alone
    queries = []
    seen = set()

    def add(q, label=""):
        if q and q not in seen:
            seen.add(q)
            queries.append((q, label))

    if name_en:
        add(f"{name_en}, {prefecture}, Greece", "en+pref")
        add(f"{name_en}, Greece", "en")
    if name_gr:
        add(f"{name_gr}, {prefecture}, Greece", "gr+pref")
        add(f"{name_gr}, Greece", "gr")
    if municipality_en and municipality_en != name_en:
        add(f"{municipality_en}, {prefecture}, Greece", "mun_en+pref")
    if municipality_gr and municipality_gr != name_gr:
        add(f"{municipality_gr}, Greece", "mun_gr")
    if municipality_gr and prefecture:
        add(f"{municipality_gr}, {prefecture}, Greece", "mun_gr+pref")

    # Try converting Greek municipality from genitive to nominative
    if municipality_gr:
        nominative = _greek_genitive_to_nominative(municipality_gr)
        if nominative and nominative != municipality_gr:
            add(f"{nominative}, Greece", "mun_nom")
            add(f"{nominative}, {prefecture}, Greece", "mun_nom+pref")

    # Try name_gr with prefecture in Greek
    if name_gr:
        prefecture_gr = _prefecture_to_greek(prefecture)
        if prefecture_gr:
            add(f"{name_gr}, {prefecture_gr}, Greece", "gr_grpref")

    for query, label in queries:
        results = _nominatim_search(query)
        if not results:
            time.sleep(GEOCODE_DELAY)
            continue

        for result in results:
            lat, lon = float(result["lat"]), float(result["lon"])
            if _validate_bounds(lat, lon, prefecture):
                return round(lat, 6), round(lon, 6), f"[{label}] {query}"

        # If first result is in Greece but outside prefecture, still accept it
        # as a last resort (but mark it)
        lat, lon = float(results[0]["lat"]), float(results[0]["lon"])
        if 34.0 <= lat <= 42.0 and 19.0 <= lon <= 30.0:
            return round(lat, 6), round(lon, 6), f"[{label}~] {query}"

        time.sleep(GEOCODE_DELAY)

    return None, None, f"[miss] {queries[0][0] if queries else name_en}"


def geocode_all_records(records: list[dict], skip_geocode: bool = False) -> list[dict]:
    """Add lat/lon to all records via multi-strategy Nominatim."""
    if skip_geocode:
        print("  [Geocode] Skipping geocoding (--skip-geocode)")
        for r in records:
            r["latitude"] = None
            r["longitude"] = None
            r["geocode_query"] = ""
        return records

    total = len(records)
    geocoded = 0
    failed_entries = []
    strategy_counts = {}

    for i, record in enumerate(records):
        name = record.get("new_name_en", "?")
        pref = record.get("prefecture", "?")
        print(f"  [Geocode] {i + 1}/{total}: {name}, {pref}...", end=" ", flush=True)

        lat, lon, query_used = geocode_settlement(record)
        record["latitude"] = lat
        record["longitude"] = lon
        record["geocode_query"] = query_used

        # Track which strategy succeeded
        strat = query_used.split("]")[0].lstrip("[") if "]" in query_used else "?"
        strategy_counts[strat] = strategy_counts.get(strat, 0) + 1

        if lat is not None:
            geocoded += 1
            print(f"({lat}, {lon}) via {strat}")
        else:
            print("NOT FOUND")
            failed_entries.append(
                {
                    "new_name_en": name,
                    "prefecture": pref,
                    "old_name_en": record.get("old_name_en", ""),
                    "new_name_gr": record.get("new_name_gr", ""),
                    "municipality_gr": record.get("municipality_gr", ""),
                    "source_url": record.get("source_url", ""),
                }
            )

        time.sleep(GEOCODE_DELAY)

    print(
        f"\n  [Geocode] Results: {geocoded}/{total} geocoded, {total - geocoded} failed"
    )
    if strategy_counts:
        print("  [Geocode] Strategy breakdown:")
        for strat, count in sorted(strategy_counts.items(), key=lambda x: -x[1]):
            print(f"    {strat}: {count}")

    # Write failed geocoding to CSV
    if failed_entries:
        with open(FAILED_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=[
                    "new_name_en",
                    "new_name_gr",
                    "prefecture",
                    "old_name_en",
                    "municipality_gr",
                    "source_url",
                ],
            )
            writer.writeheader()
            writer.writerows(failed_entries)
        print(f"  [Geocode] Failed entries written to {FAILED_FILE}")

    return records


# ---------------------------------------------------------------------------
# Phase 4: Export GeoJSON + Excel + sample
# ---------------------------------------------------------------------------


def export_geojson(records: list[dict], output_path: Path):
    """Export records as RFC 7946 GeoJSON."""
    features = []
    for r in records:
        lat = r.get("latitude")
        lon = r.get("longitude")
        if lat is None or lon is None:
            continue

        properties = {
            "old_name_gr": r.get("old_name_gr", ""),
            "old_name_en": r.get("old_name_en", ""),
            "new_name_gr": r.get("new_name_gr", ""),
            "new_name_en": r.get("new_name_en", ""),
            "prefecture": r.get("prefecture", ""),
            "province": r.get("province", ""),
            "renaming_date": r.get("renaming_date", ""),
            "settlement_code": r.get("settlement_code", ""),
            "official_journal": r.get("official_journal", ""),
            "municipality_en": r.get("municipality_en", ""),
            "municipality_gr": r.get("municipality_gr", ""),
            "source_url": r.get("source_url", ""),
        }

        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [lon, lat],  # GeoJSON is [lon, lat]
            },
            "properties": properties,
        }
        features.append(feature)

    geojson = {
        "type": "FeatureCollection",
        "name": "Greece_Settlement_Renamings",
        "crs": {
            "type": "name",
            "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"},
        },
        "features": features,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(geojson, f, ensure_ascii=False, indent=2)

    print(f"  [Export] GeoJSON: {len(features)} features → {output_path}")


def export_excel(records: list[dict], output_path: Path):
    """Export records as Excel (.xlsx) for manual auditing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Settlement Renamings"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    columns = [
        ("Old Name (GR)", 20),
        ("Old Name (EN)", 20),
        ("New Name (GR)", 20),
        ("New Name (EN)", 20),
        ("Prefecture", 18),
        ("Province", 18),
        ("Municipality (EN)", 20),
        ("Renaming Date", 14),
        ("Settlement Code", 16),
        ("Official Journal", 14),
        ("Latitude", 12),
        ("Longitude", 12),
        ("Source URL", 50),
    ]

    # Write headers
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = width

    # Fix column widths properly
    from openpyxl.utils import get_column_letter

    for col_idx, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, r in enumerate(records, 2):
        values = [
            r.get("old_name_gr", ""),
            r.get("old_name_en", ""),
            r.get("new_name_gr", ""),
            r.get("new_name_en", ""),
            r.get("prefecture", ""),
            r.get("province", ""),
            r.get("municipality_en", ""),
            r.get("renaming_date", ""),
            r.get("settlement_code", ""),
            r.get("official_journal", ""),
            r.get("latitude"),
            r.get("longitude"),
            r.get("source_url", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (11, 12) and val is not None:
                cell.number_format = "0.000000"

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  [Export] Excel: {len(records)} rows → {output_path}")


def export_sample(records: list[dict], output_path: Path, n: int = 5):
    """Export first N geocoded records as a sample JSON."""
    sample = []
    count = 0
    for r in records:
        if count >= n:
            break
        if r.get("latitude") is None:
            continue
        sample.append(
            {
                "old_name_gr": r.get("old_name_gr", ""),
                "old_name_en": r.get("old_name_en", ""),
                "new_name_gr": r.get("new_name_gr", ""),
                "new_name_en": r.get("new_name_en", ""),
                "prefecture": r.get("prefecture", ""),
                "province": r.get("province", ""),
                "renaming_date": r.get("renaming_date", ""),
                "latitude": r.get("latitude"),
                "longitude": r.get("longitude"),
            }
        )
        count += 1

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(sample, f, ensure_ascii=False, indent=2)

    print(f"  [Export] Sample: {len(sample)} rows → {output_path}")


# ---------------------------------------------------------------------------
# Progress persistence (resume support)
# ---------------------------------------------------------------------------


def save_progress(records: list[dict], entry_urls: list[dict]):
    """Save progress for resume support."""
    data = {
        "entry_urls": entry_urls,
        "records": records,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  [Progress] Saved to {PROGRESS_FILE}")


def load_progress() -> tuple[list[dict], list[dict], set]:
    """Load previous progress. Returns (records, entry_urls, scraped_urls)."""
    if not PROGRESS_FILE.exists():
        return [], [], set()

    with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    records = data.get("records", [])
    entry_urls = data.get("entry_urls", [])
    scraped_urls = {r["source_url"] for r in records if "source_url" in r}
    print(
        f"  [Progress] Loaded {len(records)} records, {len(entry_urls)} URLs from {data.get('timestamp', '?')}"
    )
    return records, entry_urls, scraped_urls


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Scrape and geocode Pandektis settlement renamings"
    )
    parser.add_argument(
        "--skip-scrape",
        action="store_true",
        help="Skip scraping, load from progress.json",
    )
    parser.add_argument(
        "--skip-geocode",
        action="store_true",
        help="Skip geocoding (useful for testing scraping)",
    )
    parser.add_argument(
        "--max-entries",
        type=int,
        default=0,
        help="Limit number of entries to scrape (0=all)",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  Greece-GEO: Pandektis Settlement Renamings Scraper")
    print("=" * 60)

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    # Phase 0: Load progress if resuming
    existing_records, entry_urls, scraped_urls = [], [], set()
    if args.skip_scrape:
        existing_records, entry_urls, scraped_urls = load_progress()
        if not entry_urls:
            print("ERROR: No progress found. Run without --skip-scrape first.")
            sys.exit(1)

    # Phase 1: Collect all entry URLs
    if not args.skip_scrape:
        print("\n[Phase 1] Collecting entry URLs from browse pages...")
        entry_urls = collect_entry_urls(session)
        print(f"  Found {len(entry_urls)} total entries")
    else:
        print(f"\n[Phase 1] Skipping (loaded {len(entry_urls)} URLs from progress)")

    # Apply max entries limit
    if args.max_entries > 0:
        entry_urls = entry_urls[: args.max_entries]
        print(f"  Limited to {args.max_entries} entries")

    # Phase 2: Parse all entry pages
    print(f"\n[Phase 2] Parsing {len(entry_urls)} entry pages...")
    if args.skip_scrape and existing_records:
        records = existing_records
        print(f"  Loaded {len(records)} existing records from progress")
    else:
        records = scrape_all_entries(session, entry_urls, skip_existing=scraped_urls)
        # Merge with existing if resuming
        if existing_records:
            records = existing_records + records
        save_progress(records, entry_urls)
    print(f"  Total records: {len(records)}")

    # Phase 3: Geocode
    print(f"\n[Phase 3] Geocoding {len(records)} records...")
    records = geocode_all_records(records, skip_geocode=args.skip_geocode)
    save_progress(records, entry_urls)

    # Phase 4: Export
    print("\n[Phase 4] Exporting files...")
    geojson_path = OUTPUT_DIR / "greece_renamings.geojson"
    excel_path = OUTPUT_DIR / "greece_renamings.xlsx"
    sample_path = OUTPUT_DIR / "sample_5_rows.json"

    export_geojson(records, geojson_path)
    export_excel(records, excel_path)
    export_sample(records, sample_path)

    # Summary
    geocoded_count = sum(1 for r in records if r.get("latitude") is not None)
    print("\n" + "=" * 60)
    print(f"  DONE: {len(records)} records scraped, {geocoded_count} geocoded")
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print(f"  Sample:  {sample_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
